"""Services against a real Postgres. Prices come from a fake provider so the tests never
depend on Yahoo being reachable."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.config import LOCAL_USER_ID as USER
from app.core.sectors import Market
from app.market.cache import PriceService
from app.services import dashboard_service, portfolio_service
from app.services.portfolio_service import UploadMode
from tests.conftest import FakeProvider

D = Decimal

INDIA_PRICES = {
    "RELIANCE": "1300",
    "HDFCBANK": "820",
    "INFY": "1100",
    "MARUTI": "13700",
    "SUNPHARMA": "1920",
    "ITC": "280",
}
US_PRICES = {"AAPL": "315", "MSFT": "385", "JPM": "336", "JNJ": "257"}


@pytest.fixture
def priced(monkeypatch):
    provider = FakeProvider({**INDIA_PRICES, **US_PRICES})
    monkeypatch.setattr(dashboard_service, "_prices", PriceService(provider))
    return provider


def edit(content: bytes, sheet: str, row: int, column: int, value) -> bytes:
    workbook = load_workbook(BytesIO(content))
    workbook[sheet].cell(row=row, column=column).value = value
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def only(content: bytes, keep_sheet: str) -> bytes:
    """Strip the holdings rows from every sheet except one."""
    workbook = load_workbook(BytesIO(content))
    for name in ("India_Holdings", "US_Holdings"):
        if name != keep_sheet:
            workbook[name].delete_rows(2, workbook[name].max_row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- Upload -------------------------------------------------------------------------


def test_bulk_upload_loads_both_markets(conn, sample_workbook):
    result = portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    assert result.ok, result.errors
    assert result.transactions_added == 10  # 6 India + 4 US
    assert result.markets == ["INDIA", "US"]


def test_incremental_upload_appends_rather_than_replacing(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    # Buy 50 more Infosys at a different price.
    more = only(edit(sample_workbook, "India_Holdings", 2, 2, "INFY"), "India_Holdings")
    more = edit(more, "India_Holdings", 2, 1, "Infosys")
    more = edit(more, "India_Holdings", 2, 3, 50)
    more = edit(more, "India_Holdings", 2, 4, 1600)
    more = edit(more, "India_Holdings", 2, 6, "IT")
    workbook = load_workbook(BytesIO(more))
    workbook["India_Holdings"].delete_rows(3, 10)  # keep only the INFY row
    buffer = BytesIO()
    workbook.save(buffer)

    result = portfolio_service.upload(conn, USER, buffer.getvalue(), UploadMode.APPEND)
    assert result.ok, result.errors

    view = dashboard_service.build(conn, USER, Market.INDIA)
    infy = next(r for r in view.stocks if r.ticker == "INFY")
    assert infy.units == D("125")  # 75 original + 50 appended
    assert infy.invested == D("170750.00")  # 75*1210 + 50*1600


def test_bulk_upload_replaces_only_the_markets_in_the_file(conn, sample_workbook):
    """An India-only re-upload must not silently wipe the US portfolio."""
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    india_only = only(sample_workbook, "India_Holdings")
    result = portfolio_service.upload(conn, USER, india_only, UploadMode.REPLACE)
    assert result.ok
    assert result.markets == ["INDIA"]

    us = dashboard_service.build(conn, USER, Market.US)
    assert us.totals.stock_count == 4  # untouched


def test_bulk_upload_replaces_within_a_market(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    view = dashboard_service.build(conn, USER, Market.INDIA)
    assert view.totals.stock_count == 6  # not 12 -- replaced, not doubled


def test_broken_file_writes_nothing(conn, sample_workbook):
    """Validate-then-commit: one bad cell rejects the whole upload."""
    broken = edit(sample_workbook, "India_Holdings", 2, 3, -5)  # negative units
    result = portfolio_service.upload(conn, USER, broken, UploadMode.REPLACE)

    assert not result.ok
    assert any(e["column"] == "Units" for e in result.errors)

    view = dashboard_service.build(conn, USER, Market.INDIA)
    assert view.totals.stock_count == 0  # the other *valid* rows did not sneak in


# --- Delete -------------------------------------------------------------------------


def test_delete_reduces_the_position(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    # The sample Deletions sheet already asks to remove 3 MARUTI of the 8 held.
    result = portfolio_service.delete_units(conn, USER, sample_workbook)
    assert result.ok, result.errors
    assert result.removed[0]["units_left"] == "5"
    assert result.removed[0]["position_closed"] is False

    view = dashboard_service.build(conn, USER, Market.INDIA)
    maruti = next(r for r in view.stocks if r.ticker == "MARUTI")
    assert maruti.units == D("5")
    assert maruti.invested == D("59000.00")  # 5 * 11800


def test_delete_all_units_removes_the_stock(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    full = edit(sample_workbook, "Deletions", 2, 3, 8)  # all 8 Maruti

    result = portfolio_service.delete_units(conn, USER, full)
    assert result.ok, result.errors
    assert result.removed[0]["position_closed"] is True

    view = dashboard_service.build(conn, USER, Market.INDIA)
    assert "MARUTI" not in {r.ticker for r in view.stocks}
    assert view.totals.stock_count == 5


def test_delete_more_than_held_is_rejected_and_writes_nothing(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    too_many = edit(sample_workbook, "Deletions", 2, 3, 500)

    result = portfolio_service.delete_units(conn, USER, too_many)
    assert not result.ok
    assert "only 8 held" in result.errors[0]["message"]

    view = dashboard_service.build(conn, USER, Market.INDIA)
    maruti = next(r for r in view.stocks if r.ticker == "MARUTI")
    assert maruti.units == D("8")  # untouched


def test_delete_unknown_ticker_is_rejected(conn, sample_workbook):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    unknown = edit(sample_workbook, "Deletions", 2, 2, "WIPRO")

    result = portfolio_service.delete_units(conn, USER, unknown)
    assert not result.ok
    assert "not in the INDIA portfolio" in result.errors[0]["message"]


def test_two_delete_rows_for_one_ticker_are_checked_against_the_combined_total(
    conn, sample_workbook
):
    """Each row passes alone (5 <= 8), but together they exceed the position."""
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    doubled = edit(sample_workbook, "Deletions", 2, 3, 5)
    doubled = edit(doubled, "Deletions", 3, 1, "INDIA")
    doubled = edit(doubled, "Deletions", 3, 2, "MARUTI")
    doubled = edit(doubled, "Deletions", 3, 3, 5)

    result = portfolio_service.delete_units(conn, USER, doubled)
    assert not result.ok
    assert "only 3 held" in result.errors[0]["message"]


# --- Dashboard ----------------------------------------------------------------------


def test_dashboard_computes_the_doc_tables(conn, sample_workbook, priced):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    view = dashboard_service.build(conn, USER, Market.INDIA)

    assert view.currency == "INR"
    assert view.totals.stock_count == 6
    assert view.unpriced == ()

    reliance = next(r for r in view.stocks if r.ticker == "RELIANCE")
    assert reliance.invested == D("59000.00")  # 50 * 1180
    assert reliance.market_value == D("65000.00")  # 50 * 1300
    assert reliance.pnl == D("6000.00")

    allocation = sum(r.allocation_pct for r in view.stocks)
    assert abs(allocation - D("100")) <= D("0.06")

    sectors = {s.sector for s in view.sectors}
    assert sectors == {
        "Energy", "Financial services", "IT", "Automobile", "Healthcare", "FMCG",
    }


def test_us_dashboard_is_independent_of_india(conn, sample_workbook, priced):
    """Each market's allocation is computed against its own total, not a combined one."""
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    us = dashboard_service.build(conn, USER, Market.US)
    assert us.currency == "USD"
    assert us.totals.stock_count == 4
    assert abs(sum(r.allocation_pct for r in us.stocks) - D("100")) <= D("0.05")

    india = dashboard_service.build(conn, USER, Market.INDIA)
    assert india.totals.invested != us.totals.invested


def test_unpriced_ticker_is_reported_not_faked(conn, sample_workbook, monkeypatch):
    """Yahoo has no price for INFY: the row shows a blank market value and INFY is
    listed as unpriced. Its allocation % is still correct, because allocation never
    touches a price."""
    partial = FakeProvider({k: v for k, v in INDIA_PRICES.items() if k != "INFY"})
    monkeypatch.setattr(dashboard_service, "_prices", PriceService(partial))

    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    view = dashboard_service.build(conn, USER, Market.INDIA)

    infy = next(r for r in view.stocks if r.ticker == "INFY")
    assert infy.market_value is None
    assert infy.pnl is None
    assert infy.allocation_pct > 0
    assert view.unpriced == ("INFY",)

    it_sector = next(s for s in view.sectors if s.sector == "IT")
    assert it_sector.unpriced_count == 1


def test_prices_are_cached_within_the_ttl(conn, sample_workbook, priced):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)

    dashboard_service.build(conn, USER, Market.INDIA)
    assert len(priced.calls) == 1

    dashboard_service.build(conn, USER, Market.INDIA)
    assert len(priced.calls) == 1  # served from price_snapshots, no second fetch

    dashboard_service.build(conn, USER, Market.INDIA, force_refresh=True)
    assert len(priced.calls) == 2  # the Refresh button skips the TTL


def test_stale_price_served_when_the_fetch_fails(conn, sample_workbook, monkeypatch):
    """A price we already have, flagged stale, beats a blank cell."""
    provider = FakeProvider(INDIA_PRICES)
    monkeypatch.setattr(dashboard_service, "_prices", PriceService(provider))
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    dashboard_service.build(conn, USER, Market.INDIA)  # warms the cache

    provider.prices = {}  # Yahoo now returns nothing at all
    view = dashboard_service.build(conn, USER, Market.INDIA, force_refresh=True)

    reliance = next(r for r in view.stocks if r.ticker == "RELIANCE")
    assert reliance.market_value == D("65000.00")  # 50 * the cached 1300
    assert reliance.stale_price is True
    assert view.unpriced == ()


# --- The agent seam -----------------------------------------------------------------


def test_snapshot_records_history_for_the_agent_layer(conn, sample_workbook, priced):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    dashboard_service.snapshot(conn, USER, Market.INDIA)

    history = dashboard_service.history(conn, USER, Market.INDIA)
    assert len(history) == 1
    row = history[0]
    assert row["stock_count"] == 6
    assert row["total_invested"] > 0
    # Sector allocations are stored so an agent can ask about drift without replaying
    # the whole ledger.
    assert set(row["sector_allocations"]) == {
        "Energy",
        "Financial services",
        "IT",
        "Automobile",
        "Healthcare",
        "FMCG",
    }


def test_snapshot_is_idempotent_within_a_day(conn, sample_workbook, priced):
    portfolio_service.upload(conn, USER, sample_workbook, UploadMode.REPLACE)
    dashboard_service.snapshot(conn, USER, Market.INDIA)
    dashboard_service.snapshot(conn, USER, Market.INDIA)

    assert len(dashboard_service.history(conn, USER, Market.INDIA)) == 1
