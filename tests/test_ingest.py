from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.sectors import Market
from app.ingest.excel_reader import read_deletions, read_holdings
from app.ingest.template_writer import build_workbook
from app.ingest.schema import HOLDINGS_SHEETS

D = Decimal


@pytest.fixture(scope="module")
def sample_bytes(tmp_path_factory) -> bytes:
    path = build_workbook(tmp_path_factory.mktemp("t") / "sample.xlsx", samples=True)
    return path.read_bytes()


def mutate(sample: bytes, sheet: str, row: int, column: int, value) -> bytes:
    workbook = load_workbook(BytesIO(sample))
    # Assign to .value rather than passing value= : openpyxl's cell() skips the
    # assignment when value is None, so clearing a cell that way silently does nothing.
    workbook[sheet].cell(row=row, column=column).value = value
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- Happy path ---------------------------------------------------------------------


def test_reads_both_markets_from_their_sheets(sample_bytes):
    rows, report = read_holdings(sample_bytes)
    assert report.ok, report.as_dicts()

    india = [r for r in rows if r.market is Market.INDIA]
    us = [r for r in rows if r.market is Market.US]
    assert len(india) == 6
    assert len(us) == 4

    reliance = next(r for r in rows if r.ticker == "RELIANCE")
    assert reliance.name == "Reliance Industries"
    assert reliance.units == D("50")
    assert reliance.sector == "Energy"
    assert reliance.purchase_date == date(2025, 4, 15)


def test_price_is_exact_decimal_not_float():
    """890.50 must survive as Decimal('890.5'), not the float approximation. This is the
    seam where float error would enter the system."""
    rows, report = read_holdings(
        build_workbook(Path("/tmp/x.xlsx"), samples=True).read_bytes()
    )
    hdfc = next(r for r in rows if r.ticker == "HDFCBANK")
    assert isinstance(hdfc.price_per_unit, Decimal)
    assert hdfc.price_per_unit == D("890.5")
    assert hdfc.units * hdfc.price_per_unit == D("89050.0")


def test_reads_deletions(sample_bytes):
    rows, report = read_deletions(sample_bytes)
    assert report.ok, report.as_dicts()
    assert len(rows) == 1
    assert rows[0].market is Market.INDIA
    assert rows[0].ticker == "MARUTI"
    assert rows[0].units == D("3")


def test_blank_template_parses_to_nothing():
    blank = build_workbook(Path("/tmp/blank.xlsx"), samples=False).read_bytes()
    rows, report = read_holdings(blank)
    assert report.ok
    assert rows == []


# --- Rejection ----------------------------------------------------------------------


def test_sector_outside_the_taxonomy_becomes_others_and_is_reported(sample_bytes):
    """The taxonomy is closed, but an unknown sector is not worth rejecting a file over.
    It becomes Others -- and is always reported, never applied silently."""
    broken = mutate(sample_bytes, "India_Holdings", 2, 6, "Crypto")
    rows, report = read_holdings(broken)

    assert report.ok  # the upload still succeeds
    assert next(r for r in rows if r.ticker == "RELIANCE").sector == "Others"

    warning = next(w for w in report.warnings if w.column == "Sector")
    assert warning.row == 2
    assert "Crypto" in warning.message
    assert "Others" in warning.message


def test_an_unknown_sector_can_be_kept_under_its_own_name(sample_bytes):
    broken = mutate(sample_bytes, "India_Holdings", 2, 6, "Renewable Energy")
    rows, report = read_holdings(broken, keep_custom_sectors=True)

    assert report.ok
    assert next(r for r in rows if r.ticker == "RELIANCE").sector == "Renewable Energy"
    assert any("kept as your own" in w.message for w in report.warnings)


def test_sectors_are_resolved_per_market(sample_bytes):
    """'Banking' means Financial services in India and Financials in the US -- the same
    idea, spelled for two different taxonomies."""
    india = mutate(sample_bytes, "India_Holdings", 2, 6, "Banking")
    rows, report = read_holdings(india)
    assert report.ok
    assert next(r for r in rows if r.ticker == "RELIANCE").sector == "Financial services"

    us = mutate(sample_bytes, "US_Holdings", 2, 6, "Banking")
    rows, report = read_holdings(us)
    assert report.ok
    assert next(r for r in rows if r.ticker == "AAPL").sector == "Financials"


def test_negative_units_rejected(sample_bytes):
    broken = mutate(sample_bytes, "India_Holdings", 3, 3, -10)
    _, report = read_holdings(broken)
    assert any(e.column == "Units" and "greater than zero" in e.message for e in report.errors)


def test_unparseable_date_rejected(sample_bytes):
    broken = mutate(sample_bytes, "India_Holdings", 4, 5, "not-a-date")
    _, report = read_holdings(broken)
    assert any(e.column == "Purchase Date" for e in report.errors)


def test_future_purchase_date_rejected(sample_bytes):
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    broken = mutate(sample_bytes, "India_Holdings", 2, 5, tomorrow)
    _, report = read_holdings(broken)
    assert any("future" in e.message for e in report.errors)


def test_missing_required_cell_rejected(sample_bytes):
    broken = mutate(sample_bytes, "India_Holdings", 2, 2, None)  # no ticker
    _, report = read_holdings(broken)
    assert any(e.column == "Ticker" and e.message == "Required." for e in report.errors)


def test_every_bad_row_is_reported_not_just_the_first(sample_bytes):
    """One upload, one complete list of problems -- so the user fixes the spreadsheet in
    a single pass instead of ten."""
    # Three genuine errors. A bad *sector* is no longer one of them -- that is a warning
    # now, and the row survives as "Others".
    broken = mutate(sample_bytes, "India_Holdings", 2, 2, "!!bad!!")  # RELIANCE: ticker
    broken = mutate(broken, "India_Holdings", 3, 3, -5)  # HDFCBANK: units
    broken = mutate(broken, "India_Holdings", 4, 5, "garbage")  # INFY: date
    rows, report = read_holdings(broken)

    assert {e.row for e in report.errors} == {2, 3, 4}
    # The good rows still parsed; the caller is what refuses to commit a partial file.
    assert {r.ticker for r in rows} == {
        "MARUTI",
        "SUNPHARMA",
        "ITC",
        "AAPL",
        "MSFT",
        "JPM",
        "JNJ",
    }


def test_missing_column_is_a_file_level_error(sample_bytes):
    # "Notes", not "Qty" -- Qty is a recognised alias for Units (brokers use it), so it
    # would resolve rather than go missing.
    broken = mutate(sample_bytes, "India_Holdings", 1, 3, "Notes")  # was 'Units'
    _, report = read_holdings(broken)
    assert any(e.row == 1 and "Missing required column" in e.message for e in report.errors)


def test_broker_headers_are_accepted_in_the_xlsx_too(sample_bytes):
    """Aliasing is a property of the reader, not of the CSV path -- a workbook whose
    columns were pasted from a broker export works just as well."""
    broker = mutate(sample_bytes, "India_Holdings", 1, 2, "Symbol")
    broker = mutate(broker, "India_Holdings", 1, 3, "Quantity Available")
    broker = mutate(broker, "India_Holdings", 1, 4, "Average Price")

    rows, report = read_holdings(broker)
    assert report.ok, report.as_dicts()
    assert next(r for r in rows if r.ticker == "RELIANCE").units == D("50")


def test_dates_accepted_in_several_formats(sample_bytes):
    for written in ("2025-04-15", "15-04-2025", "15/04/2025"):
        broken = mutate(sample_bytes, "India_Holdings", 2, 5, written)
        rows, report = read_holdings(broken)
        assert report.ok, f"{written}: {report.as_dicts()}"
        assert next(r for r in rows if r.ticker == "RELIANCE").purchase_date == date(2025, 4, 15)
