"""Builds the intake workbook.

Sector cells get a real Excel dropdown sourced from a reference sheet, so the taxonomy
is enforced at the point of typing rather than rejected an upload later. (The list is
kept on a sheet rather than inline because Excel caps an inline validation list at 255
characters and ours is longer than that.)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.core.sectors import INDIA_SECTORS, US_SECTORS, Market
from app.ingest.schema import (
    DELETIONS_COLUMNS,
    DELETIONS_SHEET,
    HOLDINGS_COLUMNS,
    HOLDINGS_SHEETS,
    SECTOR_LIST_SHEET,
)

VALIDATION_ROWS = 500  # how far down the dropdowns extend

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Real, currently-listed symbols at plausible purchase prices, so the sample portfolio
# shows a realistic mix of gains and losses against live quotes. (An earlier draft used
# TATAMOTORS, which Yahoo now 404s: Tata Motors demerged in 2025 and the passenger-
# vehicle entity trades as TMPV. Worth knowing that a ticker can simply stop existing.)
SAMPLE_INDIA = [
    ("Reliance Industries", "RELIANCE", 50, 1180.00, "2025-04-15", "Energy"),
    ("HDFC Bank", "HDFCBANK", 100, 890.50, "2025-05-20", "Financial services"),
    ("Infosys", "INFY", 75, 1210.00, "2025-06-10", "IT"),
    ("Maruti Suzuki", "MARUTI", 8, 11800.00, "2025-03-05", "Automobile"),
    ("Sun Pharmaceutical", "SUNPHARMA", 60, 1650.00, "2025-02-11", "Healthcare"),
    ("ITC", "ITC", 400, 305.00, "2025-01-18", "FMCG"),
]

SAMPLE_US = [
    ("Apple Inc", "AAPL", 25, 245.00, "2025-04-15", "Information Technology"),
    ("Microsoft Corp", "MSFT", 15, 410.00, "2025-05-20", "Information Technology"),
    ("JPMorgan Chase", "JPM", 30, 290.00, "2025-06-10", "Financials"),
    ("Johnson & Johnson", "JNJ", 40, 230.00, "2025-01-22", "Health Care"),
]


def _style_header(sheet: Worksheet, columns: tuple[str, ...], widths: list[int]) -> None:
    sheet.append(list(columns))
    for idx, width in enumerate(widths, start=1):
        cell = sheet.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.freeze_panes = "A2"


def _write_sector_lists(workbook: Workbook) -> Worksheet:
    sheet = workbook.create_sheet(SECTOR_LIST_SHEET)
    sheet["A1"], sheet["B1"], sheet["C1"] = "India Sectors", "US Sectors", "Markets"
    for i, sector in enumerate(INDIA_SECTORS, start=2):
        sheet.cell(row=i, column=1, value=sector)
    for i, sector in enumerate(US_SECTORS, start=2):
        sheet.cell(row=i, column=2, value=sector)
    for i, market in enumerate([m.value for m in Market], start=2):
        sheet.cell(row=i, column=3, value=market)
    sheet.sheet_state = "hidden"
    return sheet


def _list_ref(column: str, count: int) -> str:
    return f"'{SECTOR_LIST_SHEET}'!${column}$2:${column}${count + 1}"


def _add_dropdown(sheet: Worksheet, cells: str, source: str, prompt: str) -> None:
    validation = DataValidation(
        type="list", formula1=source, allow_blank=False, showDropDown=False
    )
    validation.error = prompt
    validation.errorTitle = "Not an allowed value"
    validation.prompt = prompt
    validation.promptTitle = "Pick from the list"
    sheet.add_data_validation(validation)
    validation.add(cells)


def _build_holdings_sheet(
    workbook: Workbook, market: Market, samples: list[tuple] | None
) -> None:
    sheet = workbook.create_sheet(HOLDINGS_SHEETS[market])
    _style_header(sheet, HOLDINGS_COLUMNS, [28, 14, 10, 16, 15, 28])

    for row in samples or []:
        sheet.append(list(row))

    for row_idx in range(2, VALIDATION_ROWS + 2):
        sheet.cell(row=row_idx, column=3).number_format = "#,##0.######"
        sheet.cell(row=row_idx, column=4).number_format = "#,##0.00"
        sheet.cell(row=row_idx, column=5).number_format = "yyyy-mm-dd"

    column, count = ("A", len(INDIA_SECTORS)) if market is Market.INDIA else ("B", len(US_SECTORS))
    _add_dropdown(
        sheet,
        f"F2:F{VALIDATION_ROWS + 1}",
        _list_ref(column, count),
        "Choose a sector from the dropdown.",
    )


def _build_deletions_sheet(workbook: Workbook, samples: bool) -> None:
    sheet = workbook.create_sheet(DELETIONS_SHEET)
    _style_header(sheet, DELETIONS_COLUMNS, [12, 14, 10])
    if samples:
        sheet.append(["INDIA", "MARUTI", 3])

    for row_idx in range(2, VALIDATION_ROWS + 2):
        sheet.cell(row=row_idx, column=3).number_format = "#,##0.######"

    _add_dropdown(
        sheet,
        f"A2:A{VALIDATION_ROWS + 1}",
        _list_ref("C", len(Market)),
        "INDIA or US.",
    )
    sheet["E1"] = (
        "Units listed here are removed from the portfolio. Remove fewer units than you "
        "hold and the position shrinks; remove all of them and the stock drops out. "
        "Units come off the oldest purchase first (FIFO)."
    )
    sheet["E1"].font = Font(italic=True, color="666666")


def build_workbook(path: Path, *, samples: bool = False) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_sector_lists(workbook)
    _build_holdings_sheet(workbook, Market.INDIA, SAMPLE_INDIA if samples else None)
    _build_holdings_sheet(workbook, Market.US, SAMPLE_US if samples else None)
    _build_deletions_sheet(workbook, samples)

    # Sector list sheet is created first so the validation refs resolve; move it last.
    workbook.move_sheet(SECTOR_LIST_SHEET, offset=len(workbook.sheetnames) - 1)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def build_all(directory: Path) -> tuple[Path, Path]:
    """A blank template to fill in, and a filled sample showing the expected format.

    Kept as two files on purpose: a template carrying example rows invites someone to
    upload Reliance and Apple by accident.
    """
    blank = build_workbook(directory / "portfolio_upload_template.xlsx", samples=False)
    sample = build_workbook(directory / "portfolio_sample.xlsx", samples=True)
    return blank, sample


if __name__ == "__main__":
    for created in build_all(Path("data/templates")):
        print(f"wrote {created}")
