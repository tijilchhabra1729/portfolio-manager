"""Parses and validates uploaded holdings -- .xlsx or .csv.

Two rules govern this module:

1. Numbers become Decimal via str(), never Decimal(float). openpyxl and csv both hand
   back values that become floats if you let them, and Decimal(2450.5) captures the
   binary approximation whereas Decimal("2450.5") is exact. This is the seam where float
   error would otherwise enter the system.
2. Every bad cell in the file is reported, not just the first. The caller then commits
   nothing unless the report is clean -- a half-applied portfolio upload is worse than a
   rejected one.

Two shapes of file arrive here:

* **Our workbook** -- two holdings sheets, and the sheet a row sits on IS its market.
* **A broker export** (Zerodha and friends) -- a flat CSV carrying the broker's own
  headers, no stock name, no purchase date, and no market. Headers are aliased onto ours;
  the market must be supplied by the caller, because nothing in the file can say.
"""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.sectors import UNCLASSIFIED, Market, resolve_sector, suggest_sectors
from app.ingest.schema import (
    COLUMN_ALIASES,
    DELETIONS_COLUMNS,
    DELETIONS_SHEET,
    HOLDINGS_COLUMNS,
    HOLDINGS_SHEETS,
    REQUIRED_HOLDINGS_COLUMNS,
    DeletionRow,
    HoldingRow,
    RowError,
    ValidationReport,
)

TICKER_RE = re.compile(r"^[A-Z0-9][A-Z0-9.\-&]{0,19}$")
DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y")

Source = Path | BytesIO | bytes


# --- headers ------------------------------------------------------------------------


def _squash(text: str) -> str:
    cleaned = "".join(c if c.isalnum() or c.isspace() else " " for c in text.lower())
    return " ".join(cleaned.split())


def _canonical(header: str) -> str:
    """Map a broker's column name onto ours: 'Quantity Available' -> 'Units'."""
    key = _squash(header)
    if key in COLUMN_ALIASES:
        return COLUMN_ALIASES[key]
    for name in HOLDINGS_COLUMNS + DELETIONS_COLUMNS:
        if key == _squash(name):
            return name
    return header.strip()


def _map_headers(names: list[Any]) -> dict[str, int]:
    return {
        _canonical(str(name)): idx
        for idx, name in enumerate(names)
        if name is not None and str(name).strip()
    }


def _missing(
    columns: dict[str, int],
    required: tuple[str, ...],
    sheet: str,
    errors: list[RowError],
) -> bool:
    absent = [c for c in required if c not in columns]
    if absent:
        errors.append(
            RowError(sheet, 1, ", ".join(absent), "Missing required column(s).")
        )
    return bool(absent)


def _is_blank(values: tuple) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


# --- cell parsing --------------------------------------------------------------------


class _RowParser:
    """Reads one row, appending to the shared error list instead of raising, so a single
    pass surfaces every problem in the file."""

    def __init__(
        self,
        sheet: str,
        row: int,
        errors: list[RowError],
        warnings: list[RowError] | None = None,
        keep_custom_sectors: bool = False,
    ) -> None:
        self.sheet = sheet
        self.row = row
        self.errors = errors
        self.warnings = warnings if warnings is not None else []
        self.keep_custom_sectors = keep_custom_sectors
        self.failed = False

    def _fail(self, column: str, message: str) -> None:
        self.errors.append(RowError(self.sheet, self.row, column, message))
        self.failed = True

    def _warn(self, column: str, message: str) -> None:
        self.warnings.append(RowError(self.sheet, self.row, column, message))

    def text(self, column: str, value: Any, required: bool = True) -> str:
        if value is None or not str(value).strip():
            if required:
                self._fail(column, "Required.")
            return ""
        return str(value).strip()

    def ticker(self, column: str, value: Any) -> str:
        raw = self.text(column, value)
        if not raw:
            return ""
        symbol = raw.upper().replace(" ", "")
        if not TICKER_RE.match(symbol):
            self._fail(column, f"'{raw}' is not a valid ticker symbol.")
            return ""
        return symbol

    def positive_decimal(self, column: str, value: Any) -> Decimal:
        if value is None or (isinstance(value, str) and not value.strip()):
            self._fail(column, "Required.")
            return Decimal(0)
        try:
            # str() first: Decimal(2450.5) would bake in the float approximation.
            number = Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            self._fail(column, f"'{value}' is not a number.")
            return Decimal(0)
        if number <= 0:
            self._fail(column, "Must be greater than zero.")
            return Decimal(0)
        return number

    def past_date(self, column: str, value: Any) -> date:
        """Optional. A broker export reports an already-averaged position rather than a
        lot history, so there is no purchase date to give -- we date it today instead of
        making the user invent one. Nothing is lost: with one averaged lot per stock,
        FIFO has nothing left to order."""
        if value is None or not str(value).strip():
            return date.today()

        parsed: date | None = None
        if isinstance(value, datetime):
            parsed = value.date()
        elif isinstance(value, date):
            parsed = value
        else:
            raw = str(value).strip()
            for fmt in DATE_FORMATS:
                try:
                    parsed = datetime.strptime(raw, fmt).date()
                    break
                except ValueError:
                    continue

        if parsed is None:
            self._fail(column, f"'{value}' is not a date (use YYYY-MM-DD).")
            return date.today()
        if parsed > date.today():
            self._fail(column, "Purchase date is in the future.")
        return parsed

    def sector(self, column: str, value: Any, market: Market) -> str:
        """The taxonomy is closed. An unrecognised sector is not a reason to reject the
        row -- it becomes "Others", or the user's own name if they asked for that -- but
        it is always reported. Silently reclassifying a holding is how a sector
        allocation goes wrong without anyone noticing."""
        raw = self.text(column, value, required=False)
        if not raw:
            self._warn(column, f"No sector given; filed under {UNCLASSIFIED}.")
            return UNCLASSIFIED

        resolved = resolve_sector(market, raw)
        if resolved is not None:
            return resolved

        if self.keep_custom_sectors:
            custom = " ".join(raw.split())[:64]
            self._warn(column, f"'{custom}' is not a known sector; kept as your own.")
            return custom

        hint = suggest_sectors(market, raw)
        suffix = f" Did you mean {' or '.join(hint)}?" if hint else ""
        self._warn(
            column,
            f"'{raw}' is not a known {market.value} sector; filed under "
            f"{UNCLASSIFIED}.{suffix}",
        )
        return UNCLASSIFIED

    def market(self, column: str, value: Any) -> Market | None:
        raw = self.text(column, value)
        if not raw:
            return None
        try:
            return Market(raw.upper())
        except ValueError:
            self._fail(column, f"'{raw}' is not a market. Use INDIA or US.")
            return None


# --- row assembly --------------------------------------------------------------------


def _read_rows(
    rows: list[tuple],
    columns: dict[str, int],
    market: Market,
    sheet: str,
    errors: list[RowError],
    warnings: list[RowError],
    keep_custom_sectors: bool = False,
) -> list[HoldingRow]:
    out: list[HoldingRow] = []
    for row_idx, values in enumerate(rows, start=2):
        if _is_blank(values):
            continue

        def cell(name: str) -> Any:
            idx = columns.get(name)
            return values[idx] if idx is not None and idx < len(values) else None

        parser = _RowParser(sheet, row_idx, errors, warnings, keep_custom_sectors)
        ticker = parser.ticker("Ticker", cell("Ticker"))
        row = HoldingRow(
            market=market,
            # A broker export carries no company name. Fall back to the symbol rather
            # than rejecting the row over a cosmetic field.
            name=parser.text("Stock Name", cell("Stock Name"), required=False) or ticker,
            ticker=ticker,
            units=parser.positive_decimal("Units", cell("Units")),
            price_per_unit=parser.positive_decimal(
                "Price Per Unit", cell("Price Per Unit")
            ),
            purchase_date=parser.past_date("Purchase Date", cell("Purchase Date")),
            sector=parser.sector("Sector", cell("Sector"), market),
        )
        if not parser.failed:
            out.append(row)
    return out


# --- entry points --------------------------------------------------------------------


def _as_bytes(source: Source) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, Path):
        return source.read_bytes()
    return source.getvalue()


def is_csv(filename: str | None) -> bool:
    return bool(filename) and filename.lower().endswith(".csv")


def read_holdings(
    source: Source,
    *,
    filename: str | None = None,
    market: Market | None = None,
    keep_custom_sectors: bool = False,
) -> tuple[list[HoldingRow], ValidationReport]:
    """Read a holdings file.

    `market` is required for a CSV -- a flat file names no market anywhere, so the upload
    form has to say which one it is. For our workbook it is ignored: the sheet a row sits
    on already decides.

    `keep_custom_sectors` keeps an unrecognised sector under its own name instead of
    filing it under "Others". Off by default, because a typo left to itself becomes an
    industry.
    """
    if is_csv(filename):
        return _read_csv(_as_bytes(source), market, keep_custom_sectors)
    return _read_workbook(_as_bytes(source), keep_custom_sectors)


def _read_csv(
    content: bytes, market: Market | None, keep_custom_sectors: bool = False
) -> tuple[list[HoldingRow], ValidationReport]:
    errors: list[RowError] = []
    warnings: list[RowError] = []
    if market is None:
        errors.append(
            RowError(
                "(file)", 0, "market",
                "A CSV does not say which market it belongs to. Choose India or US on "
                "the upload form.",
            )
        )
        return [], ValidationReport(errors, warnings)

    # utf-8-sig: Excel prefixes a BOM when it saves a CSV, which would otherwise glue
    # itself to the first header and make 'Symbol' unrecognisable.
    rows = list(csv.reader(content.decode("utf-8-sig", errors="replace").splitlines()))
    if not rows:
        errors.append(RowError("(file)", 0, "-", "The file is empty."))
        return [], ValidationReport(errors, warnings)

    columns = _map_headers(list(rows[0]))
    if _missing(columns, REQUIRED_HOLDINGS_COLUMNS, "CSV", errors):
        return [], ValidationReport(errors)

    parsed = _read_rows(
        [tuple(r) for r in rows[1:]], columns, market, "CSV", errors, warnings,
        keep_custom_sectors,
    )
    return parsed, ValidationReport(errors, warnings)


def _read_workbook(
    content: bytes, keep_custom_sectors: bool = False
) -> tuple[list[HoldingRow], ValidationReport]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    errors: list[RowError] = []
    warnings: list[RowError] = []
    parsed: list[HoldingRow] = []
    found = False

    for market, sheet_name in HOLDINGS_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        found = True
        sheet = workbook[sheet_name]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        columns = _map_headers(list(header))
        if _missing(columns, REQUIRED_HOLDINGS_COLUMNS, sheet_name, errors):
            continue

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        parsed += _read_rows(
            rows, columns, market, sheet_name, errors, warnings, keep_custom_sectors
        )

    if not found:
        expected = " or ".join(HOLDINGS_SHEETS.values())
        errors.append(
            RowError(
                "(workbook)", 0, "sheets",
                f"No {expected} sheet found. If this is a broker export, upload it as "
                ".csv and pick the market on the upload form.",
            )
        )

    return parsed, ValidationReport(errors, warnings)


def read_deletions(source: Source) -> tuple[list[DeletionRow], ValidationReport]:
    workbook = load_workbook(BytesIO(_as_bytes(source)), data_only=True)
    errors: list[RowError] = []
    rows: list[DeletionRow] = []

    if DELETIONS_SHEET not in workbook.sheetnames:
        errors.append(
            RowError("(workbook)", 0, "sheets", f"No {DELETIONS_SHEET} sheet found.")
        )
        return rows, ValidationReport(errors)

    sheet = workbook[DELETIONS_SHEET]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = _map_headers(list(header))
    if _missing(columns, DELETIONS_COLUMNS, DELETIONS_SHEET, errors):
        return rows, ValidationReport(errors)

    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # The Deletions sheet carries a help note in a further column; judge emptiness on
        # the three real columns only.
        if _is_blank(tuple(values[i] for i in columns.values() if i < len(values))):
            continue

        def cell(name: str) -> Any:
            idx = columns[name]
            return values[idx] if idx < len(values) else None

        parser = _RowParser(DELETIONS_SHEET, row_idx, errors)
        market = parser.market("Market", cell("Market"))
        ticker = parser.ticker("Ticker", cell("Ticker"))
        units = parser.positive_decimal("Units", cell("Units"))
        if not parser.failed and market is not None:
            rows.append(DeletionRow(market=market, ticker=ticker, units=units))

    return rows, ValidationReport(errors)
