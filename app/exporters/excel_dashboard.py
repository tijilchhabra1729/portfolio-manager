"""The dashboard as a spreadsheet.

The doc asked for the dashboard to be an Excel output with charts alongside the tables.
The website supersedes it as the primary surface, but the export is still here -- it is
what you attach to an email or hand to someone who wants the numbers in a pivot table.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.models import DashboardView

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
GAIN = Font(color="1B7F4B")
LOSS = Font(color="C0392B")

STOCK_COLUMNS = (
    "Sno",
    "Stock Name",
    "Ticker",
    "Sector",
    "Total Units",
    "Invested Amount",
    "Market Value",
    "Profit/Loss",
    "Profit/Loss %",
    "Allocation %",
)
SECTOR_COLUMNS = (
    "Sno",
    "Sector Name",
    "Stocks",
    "Invested Amount",
    "Market Value",
    "Profit/Loss",
    "Profit/Loss %",
    "Allocation %",
)


def _num(value: Decimal | None) -> float | str:
    """Excel cells hold floats. Every calculation is already finished and rounded by the
    time it reaches here, so this converts for display only -- no arithmetic follows."""
    return "-" if value is None else float(value)


def _write_header(sheet: Worksheet, row: int, columns: tuple[str, ...]) -> None:
    for idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _colour_pnl(sheet: Worksheet, row: int, columns: tuple[int, ...]) -> None:
    for col in columns:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell.value, float):
            cell.font = GAIN if cell.value >= 0 else LOSS


def _stocks_sheet(workbook: Workbook, view: DashboardView) -> Worksheet:
    sheet = workbook.create_sheet("Stock Allocation")
    money = f'#,##0.00" {view.currency}"'

    sheet["A1"] = f"Stock Allocation — {view.market.value}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"As of {view.generated_at:%Y-%m-%d %H:%M UTC}"
    sheet["A2"].font = Font(italic=True, color="666666")

    _write_header(sheet, 4, STOCK_COLUMNS)
    for offset, row in enumerate(view.stocks):
        r = 5 + offset
        values = (
            row.sno,
            row.name,
            row.ticker,
            row.sector,
            _num(row.units),
            _num(row.invested),
            _num(row.market_value),
            _num(row.pnl),
            _num(row.pnl_pct),
            _num(row.allocation_pct),
        )
        for idx, value in enumerate(values, start=1):
            sheet.cell(row=r, column=idx, value=value)
        for col in (6, 7, 8):
            sheet.cell(row=r, column=col).number_format = money
        for col in (9, 10):
            sheet.cell(row=r, column=col).number_format = '0.00"%"'
        _colour_pnl(sheet, r, (8, 9))

    total_row = 5 + len(view.stocks)
    sheet.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    for col, value in (
        (6, view.totals.invested),
        (7, view.totals.market_value),
        (8, view.totals.pnl),
    ):
        cell = sheet.cell(row=total_row, column=col, value=_num(value))
        cell.font = Font(bold=True)
        cell.number_format = money

    for idx, width in enumerate((6, 26, 14, 24, 12, 18, 18, 16, 14, 13), start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.freeze_panes = "A5"
    return sheet


def _sectors_sheet(workbook: Workbook, view: DashboardView) -> Worksheet:
    sheet = workbook.create_sheet("Sector Allocation")
    money = f'#,##0.00" {view.currency}"'

    sheet["A1"] = f"Sector Allocation — {view.market.value}"
    sheet["A1"].font = TITLE_FONT

    _write_header(sheet, 3, SECTOR_COLUMNS)
    for offset, row in enumerate(view.sectors):
        r = 4 + offset
        values = (
            row.sno,
            row.sector,
            row.stock_count,
            _num(row.invested),
            _num(row.market_value),
            _num(row.pnl),
            _num(row.pnl_pct),
            _num(row.allocation_pct),
        )
        for idx, value in enumerate(values, start=1):
            sheet.cell(row=r, column=idx, value=value)
        for col in (4, 5, 6):
            sheet.cell(row=r, column=col).number_format = money
        for col in (7, 8):
            sheet.cell(row=r, column=col).number_format = '0.00"%"'
        _colour_pnl(sheet, r, (6, 7))

    for idx, width in enumerate((6, 28, 9, 18, 18, 16, 14, 13), start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.freeze_panes = "A4"
    return sheet


def _charts(workbook: Workbook, view: DashboardView) -> None:
    if not view.sectors:
        return
    sheet = workbook.create_sheet("Charts")
    last = 3 + len(view.sectors)

    pie = PieChart()
    pie.title = "Sector allocation (% of invested amount)"
    pie.height, pie.width = 10, 16
    pie.add_data(
        Reference(workbook["Sector Allocation"], min_col=8, min_row=3, max_row=last),
        titles_from_data=True,
    )
    pie.set_categories(
        Reference(workbook["Sector Allocation"], min_col=2, min_row=4, max_row=last)
    )
    sheet.add_chart(pie, "B2")

    bar = BarChart()
    bar.type, bar.style = "col", 10
    bar.title = "Profit / loss by sector"
    bar.y_axis.title = view.currency
    bar.height, bar.width = 10, 16
    bar.add_data(
        Reference(workbook["Sector Allocation"], min_col=6, min_row=3, max_row=last),
        titles_from_data=True,
    )
    bar.set_categories(
        Reference(workbook["Sector Allocation"], min_col=2, min_row=4, max_row=last)
    )
    sheet.add_chart(bar, "B24")


def export(view: DashboardView) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Summary")
    summary["A1"] = f"Portfolio — {view.market.value}"
    summary["A1"].font = TITLE_FONT
    rows = (
        ("Generated", f"{view.generated_at:%Y-%m-%d %H:%M UTC}"),
        ("Currency", view.currency),
        ("Holdings", view.totals.stock_count),
        ("Sectors", view.totals.sector_count),
        ("Invested amount", _num(view.totals.invested)),
        ("Market value", _num(view.totals.market_value)),
        ("Profit / loss", _num(view.totals.pnl)),
        ("Profit / loss %", _num(view.totals.pnl_pct)),
    )
    for offset, (label, value) in enumerate(rows, start=3):
        summary.cell(row=offset, column=1, value=label).font = Font(bold=True)
        summary.cell(row=offset, column=2, value=value)
    if view.unpriced:
        summary.cell(
            row=3 + len(rows) + 1,
            column=1,
            value=f"Could not price: {', '.join(view.unpriced)}",
        ).font = Font(italic=True, color="C0392B")
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 28

    _stocks_sheet(workbook, view)
    _sectors_sheet(workbook, view)
    _charts(workbook, view)
    return workbook
